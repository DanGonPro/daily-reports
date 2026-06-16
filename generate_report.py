import sqlite3
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

DB_PATH = os.path.join(os.path.dirname(__file__), "traffic.db")
REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")

def get_conn():
    return sqlite3.connect(DB_PATH)

def style_header(ws, row, cols, fill_color="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style="medium", color="FFFFFF")
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

def build_report():
    os.makedirs(REPORTS_DIR, exist_ok=True)
    conn = get_conn()
    cur = conn.cursor()
    today = datetime.today()
    today_str = today.strftime("%Y-%m-%d")
    yesterday = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    last_7 = (today - timedelta(days=7)).strftime("%Y-%m-%d")
    last_30 = (today - timedelta(days=30)).strftime("%Y-%m-%d")

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Website Traffic Report"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Generated: {today.strftime('%B %d, %Y %I:%M %p')}"
    ws["A2"].font = Font(italic=True, color="595959", size=10)

    headers = ["Period", "Total Sessions", "Total Page Views", "Avg Bounce Rate", "Avg Duration (sec)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i).value = h
    style_header(ws, 4, len(headers))

    periods = [
        ("Yesterday", yesterday, yesterday),
        ("Last 7 Days", last_7, today_str),
        ("Last 30 Days", last_30, today_str),
    ]
    for row_idx, (label, date_from, date_to) in enumerate(periods, start=5):
        cur.execute("""
            SELECT SUM(sessions), SUM(page_views), AVG(bounce_rate), AVG(avg_duration_sec)
            FROM traffic WHERE date BETWEEN ? AND ?
        """, (date_from, date_to))
        r = cur.fetchone()
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=2).value = int(r[0] or 0)
        ws.cell(row=row_idx, column=3).value = int(r[1] or 0)
        ws.cell(row=row_idx, column=4).value = f"{round((r[2] or 0)*100, 1)}%"
        ws.cell(row=row_idx, column=5).value = int(r[3] or 0)
        fill = PatternFill("solid", fgColor="D6E4F0" if row_idx % 2 == 0 else "EBF5FB")
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = fill
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

    auto_width(ws)

    # ── Sheet 2: Top Pages (Last 7 Days) ─────────────────────────────────────
    ws2 = wb.create_sheet("Top Pages")
    ws2["A1"] = "Top Pages — Last 7 Days"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")

    page_headers = ["Page", "Sessions", "Page Views", "Avg Bounce Rate", "Avg Duration (sec)"]
    for i, h in enumerate(page_headers, 1):
        ws2.cell(row=3, column=i).value = h
    style_header(ws2, 3, len(page_headers))

    cur.execute("""
        SELECT page, SUM(sessions), SUM(page_views), AVG(bounce_rate), AVG(avg_duration_sec)
        FROM traffic WHERE date BETWEEN ? AND ?
        GROUP BY page ORDER BY SUM(sessions) DESC
    """, (last_7, today_str))

    for row_idx, row in enumerate(cur.fetchall(), start=4):
        ws2.cell(row=row_idx, column=1).value = row[0]
        ws2.cell(row=row_idx, column=2).value = int(row[1])
        ws2.cell(row=row_idx, column=3).value = int(row[2])
        ws2.cell(row=row_idx, column=4).value = f"{round(row[3]*100, 1)}%"
        ws2.cell(row=row_idx, column=5).value = int(row[4])
        fill = PatternFill("solid", fgColor="D6E4F0" if row_idx % 2 == 0 else "EBF5FB")
        for col in range(1, 6):
            ws2.cell(row=row_idx, column=col).fill = fill
            ws2.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

    auto_width(ws2)

    # ── Sheet 3: Traffic by Country ───────────────────────────────────────────
    ws3 = wb.create_sheet("By Country")
    ws3["A1"] = "Traffic by Country — Last 7 Days"
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E79")

    country_headers = ["Country", "Sessions", "Page Views", "Avg Bounce Rate"]
    for i, h in enumerate(country_headers, 1):
        ws3.cell(row=3, column=i).value = h
    style_header(ws3, 3, len(country_headers))

    cur.execute("""
        SELECT country, SUM(sessions), SUM(page_views), AVG(bounce_rate)
        FROM traffic WHERE date BETWEEN ? AND ?
        GROUP BY country ORDER BY SUM(sessions) DESC
    """, (last_7, today_str))

    for row_idx, row in enumerate(cur.fetchall(), start=4):
        ws3.cell(row=row_idx, column=1).value = row[0]
        ws3.cell(row=row_idx, column=2).value = int(row[1])
        ws3.cell(row=row_idx, column=3).value = int(row[2])
        ws3.cell(row=row_idx, column=4).value = f"{round(row[3]*100, 1)}%"
        fill = PatternFill("solid", fgColor="D6E4F0" if row_idx % 2 == 0 else "EBF5FB")
        for col in range(1, 5):
            ws3.cell(row=row_idx, column=col).fill = fill
            ws3.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

    auto_width(ws3)

    # ── Sheet 4: Traffic by Source ────────────────────────────────────────────
    ws4 = wb.create_sheet("By Source")
    ws4["A1"] = "Traffic by Source — Last 7 Days"
    ws4["A1"].font = Font(bold=True, size=14, color="1F4E79")

    source_headers = ["Source", "Sessions", "Page Views", "Avg Bounce Rate"]
    for i, h in enumerate(source_headers, 1):
        ws4.cell(row=3, column=i).value = h
    style_header(ws4, 3, len(source_headers))

    cur.execute("""
        SELECT source, SUM(sessions), SUM(page_views), AVG(bounce_rate)
        FROM traffic WHERE date BETWEEN ? AND ?
        GROUP BY source ORDER BY SUM(sessions) DESC
    """, (last_7, today_str))

    for row_idx, row in enumerate(cur.fetchall(), start=4):
        ws4.cell(row=row_idx, column=1).value = row[0]
        ws4.cell(row=row_idx, column=2).value = int(row[1])
        ws4.cell(row=row_idx, column=3).value = int(row[2])
        ws4.cell(row=row_idx, column=4).value = f"{round(row[3]*100, 1)}%"
        fill = PatternFill("solid", fgColor="D6E4F0" if row_idx % 2 == 0 else "EBF5FB")
        for col in range(1, 5):
            ws4.cell(row=row_idx, column=col).fill = fill
            ws4.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

    auto_width(ws4)

    conn.close()

    timestamp = today.strftime("%Y-%m-%d_%H-%M")
    filename = f"traffic_report_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)
    print(f"Report saved: {filepath}")

if __name__ == "__main__":
    build_report()
